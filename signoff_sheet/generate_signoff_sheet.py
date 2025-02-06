import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar
import json
from typing import Dict, List
import os
import argparse
from datetime import datetime

# Get the directory where this script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def save_default_configs():
    """Save the default configurations as JSON files."""
    production_config = {
        "Production": {
            "Weekly": {
                "description": "Sweep Floor, Clean walls/windows/door frames",
                "columns": ["1.94,1.97", "1.95", "1.18,1.20,1.50", "8.17", "8.51,8.73", "8.7"]
            },
            "Quarterly": {
                "description": "Scrub Floors",
                "columns": ["1.94,1.97", "1.95", "1.18,1.20,1.50", "8.17", "8.51,8.73", "8.7"]
            }
        }
    }
    
    warehouse_config = {
        "Warehouse": {
            "Daily": {
                "description": "Dust mop floor, empty trash/replace liners, clean corners, wipe down walls/windows/door frames, change dust mop head weekly, inspect walls/windows for damage",
                "columns": ["8.72","8.71,8.46,8.47","7.08, 7.09, 7.17, 7.19, 7.20","8.77A, 8.77B","7.00-7.03", "1.58","1.02,1.08,5.29"]
            },
            "2x Weekly": {
                "description": "Scrub/Mop Floor",
                "columns": ["8.72","8.71,8.46,8.47","7.08, 7.09, 7.17, 7.19, 7.20","8.77A, 8.77B","7.00-7.03", "1.58","1.02,1.08,5.29"]
            },
            "Monthly": {
                "description": "Deep clean storage",
                "columns": ["8.72","8.71,8.46,8.47","7.08, 7.09, 7.17, 7.19, 7.20","8.77A, 8.77B","7.00-7.03", "1.58","1.02,1.08,5.29"]
            }
        }
    }

    idf_config = {
        "IDF/Server Room": {
            "Monthly": {
                "description": "Dust and sweep/mop floors",
                "columns": ["9.93","9.71","9.48","9.44","9.20","9.05","8.74","8.68","8.59","8.33","8.07","7.14","7.04","6.54","5.04","4.01","2.38","2.24,2.25","1.92","1.68","1.53","0.42,0.43","0.97"]
            }
        }
    }
    
    # Create configs directory if it doesn't exist
    configs_dir = os.path.join(SCRIPT_DIR, 'configs')
    os.makedirs(configs_dir, exist_ok=True)
    
    # Save configurations
    with open(os.path.join(configs_dir, 'production_config.json'), 'w') as f:
        json.dump(production_config, f, indent=4)
    
    with open(os.path.join(configs_dir, 'warehouse_config.json'), 'w') as f:
        json.dump(warehouse_config, f, indent=4)

    with open(os.path.join(configs_dir, 'idf_config.json'), 'w') as f:
        json.dump(idf_config, f, indent=4)

def load_config(config_file: str) -> dict:
    """
    Load configuration from a JSON file.
    
    Args:
        config_file (str): Path to the configuration JSON file
    
    Returns:
        dict: Configuration dictionary
    """
    config_path = os.path.join(SCRIPT_DIR, config_file)
    with open(config_path, 'r') as f:
        return json.load(f)

def get_month_year_input():
    """
    Get month and year input from the user if not provided via command line.
    
    Returns:
        tuple: (month, year)
    """
    while True:
        try:
            date_str = input("Enter month and year (MM/YYYY): ")
            date = datetime.strptime(date_str, "%m/%Y")
            return date.month, date.year
        except ValueError:
            print("Invalid format. Please use MM/YYYY (e.g., 01/2025)")

def create_task_report(
    month: int,
    year: int,
    output_filename: str,
    config: Dict[str, Dict[str, List[str]]]
):
    """
    Create a task report based on provided configuration.
    
    Args:
        month (int): Month number (1-12)
        year (int): Year
        output_filename (str): Name of the output Excel file
        config (dict): Configuration dictionary loaded from JSON file
    """
    # Create a new workbook
    wb = Workbook()
    
    # Define styles
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center')
    
    # Process each frequency in a separate sheet
    first_sheet = True
    for section_name, frequencies in config.items():
        for freq_name, freq_config in frequencies.items():
            # Create a new sheet for each frequency
            if first_sheet:
                ws = wb.active
                ws.title = freq_name.replace('/', '_')  # Replace invalid characters
                first_sheet = False
            else:
                ws = wb.create_sheet(freq_name.replace('/', '_'))  # Replace invalid characters
            
            # Add section and frequency headers
            ws.merge_cells('A1:A2')
            ws['A1'] = 'Date'
            ws['A1'].alignment = center_aligned
            ws['A1'].font = Font(bold=True)
            ws['A1'].border = border
            
            # Add description header
            description_text = f"{section_name} - {freq_name}\n({freq_config['description']})"
            ws.merge_cells(f'B1:H1')  # Merge cells for the description
            ws['B1'] = description_text
            ws['B1'].alignment = center_aligned
            ws['B1'].font = Font(bold=True)
            ws['B1'].border = border
            
            # Add room number headers
            for i, room in enumerate(freq_config['columns'], start=2):  # Start from column B
                col = get_column_letter(i)
                ws[f'{col}2'] = room
                ws[f'{col}2'].alignment = center_aligned
                ws[f'{col}2'].font = Font(bold=True)
                ws[f'{col}2'].border = border
            
            # Generate dates for the month
            num_days = calendar.monthrange(year, month)[1]
            start_date = datetime(year, month, 1)
            dates = [(start_date + timedelta(days=i)).strftime('%a, %b %-d %Y') 
                     for i in range(num_days)]
            
            # Add dates and empty cells
            for i, date in enumerate(dates, start=3):  # Start from row 3 (after headers)
                # Get the day of the week (0 = Monday, 6 = Sunday)
                current_date = start_date + timedelta(days=i-3)
                day_of_week = current_date.weekday()
                
                # Create thicker bottom border for end of week (Sunday)
                current_border = border
                if day_of_week == 6:  # Sunday
                    current_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium')  # Thicker bottom border
                    )
                
                # Add date
                ws[f'A{i}'] = date
                ws[f'A{i}'].alignment = left_aligned
                ws[f'A{i}'].border = current_border
                
                # Add empty cells for all columns
                for col in range(2, len(freq_config['columns']) + 2):
                    cell = ws.cell(row=i, column=col)
                    cell.border = current_border
                    cell.alignment = center_aligned
            
            # Add "Reviewed by" section
            review_row = len(dates) + 5
            ws[f'A{review_row}'] = 'Reviewed by'
            ws[f'A{review_row}'].font = Font(bold=True)
            
            # Add signature lines
            fields = ['Sign:', 'Name:', 'Date:']
            for i, field in enumerate(fields):
                current_row = review_row + i + 2
                ws[f'A{current_row}'] = field
                ws.merge_cells(f'B{current_row}:D{current_row}')
                ws[f'B{current_row}'].border = Border(bottom=Side(style='thin'))
            
            # Set column widths
            ws.column_dimensions['A'].width = 20
            for col in range(2, len(freq_config['columns']) + 2):
                ws.column_dimensions[get_column_letter(col)].width = 12
            
            # Set row heights for headers
            ws.row_dimensions[1].height = 30  # Description row
            ws.row_dimensions[2].height = 30  # Room numbers
    
    # Save the workbook
    wb.save(output_filename)

def normalize_string(s: str) -> str:
    """
    Normalize a string by converting to lowercase and standardizing spaces.
    Special handling for:
    - Frequencies with '2x' or 'twice'
    - Room numbers with commas or letters
    - Multiple rooms in a single string
    
    Args:
        s (str): String to normalize
    
    Returns:
        str: Normalized string
    """
    # Convert to lowercase and strip whitespace
    s = s.lower().strip()
    
    # Special handling for frequencies with '2x' or 'twice'
    if '2x ' in s:
        parts = s.split('2x ')
        return '2x ' + ' '.join(parts[1].split())
    elif 'twice ' in s:
        parts = s.split('twice ')
        return 'twice ' + ' '.join(parts[1].split())
    
    # Special handling for room numbers
    if any(c.isdigit() for c in s):  # If string contains numbers (likely a room number)
        # Remove spaces around commas and preserve letters after numbers
        parts = [part.strip() for part in s.split(',')]
        normalized_parts = []
        for part in parts:
            # Keep letters that are directly after numbers (e.g., '8.77a')
            # but normalize any other spaces
            cleaned = ' '.join(part.split())
            normalized_parts.append(cleaned)
        return ','.join(normalized_parts)
    
    # For other cases (like frequencies without '2x'), normalize spaces
    return ' '.join(s.split())

def write_unmatched_tasks(unmatched_records: list, excel_file: str, month: int, year: int):
    """Write unmatched tasks to a review file"""
    review_dir = os.path.join(SCRIPT_DIR, 'data', 'review')
    os.makedirs(review_dir, exist_ok=True)
    
    config_type = os.path.basename(excel_file).split('_')[0]  # Get Production/Warehouse/Idf part
    review_file = os.path.join(review_dir, f'unmatched_tasks_{config_type}_{year}_{month:02d}.txt')
    
    with open(review_file, 'w') as f:
        f.write(f"Unmatched Tasks Review for {config_type} - {calendar.month_name[month]} {year}\n")
        f.write("=" * 80 + "\n\n")
        
        # Group by frequency and room number
        from collections import defaultdict
        grouped_errors = defaultdict(int)
        for error in unmatched_records:
            if "No matching column for frequency" in error:
                # Extract frequency and room from error message
                parts = error.split("'")
                if len(parts) >= 4:
                    freq = parts[1]
                    room = parts[3]
                    grouped_errors[(freq, room)] += 1
        
        # Write summary of unmatched combinations
        f.write("Summary of Unmatched Combinations:\n")
        f.write("-" * 40 + "\n")
        for (freq, room), count in sorted(grouped_errors.items()):
            f.write(f"Frequency: '{freq}', Room: '{room}' - {count} occurrences\n")
        
        # Write all individual errors
        f.write("\n\nDetailed Error List:\n")
        f.write("-" * 40 + "\n")
        for error in unmatched_records:
            f.write(f"{error}\n")
    
    print(f"\nUnmatched tasks written to: {review_file}")

def populate_report_data(excel_file: str, completion_data_csv: str, month: int, year: int):
    """
    Populate an existing task report template with completion data from a CSV file.
    
    Args:
        excel_file (str): Path to the existing Excel report template
        completion_data_csv (str): Path to the CSV file containing completion data
        month (int): Month number (1-12)
        year (int): Year (e.g., 2025)
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    import os.path
    
    # Create target date object
    target_date = datetime(year, month, 1)
    
    # Read the completion data
    try:
        df = pd.read_csv(completion_data_csv)
    except FileNotFoundError:
        raise FileNotFoundError(f"Completion data file not found: {os.path.basename(completion_data_csv)}")
    except pd.errors.EmptyDataError:
        raise ValueError(f"Completion data file is empty: {os.path.basename(completion_data_csv)}")
    except Exception as e:
        raise ValueError(f"Error reading completion data file: {str(e)}")
    
    # Validate required columns
    required_columns = ['Datetime', 'Name', 'Room Number', 'Frequency', 'Task Type']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns in CSV: {', '.join(missing_columns)}")
    
    # Convert datetime to proper format
    try:
        df['Datetime'] = pd.to_datetime(df['Datetime'])
    except Exception as e:
        raise ValueError(f"Error converting datetime column: {str(e)}")
    
    # Get the configuration type from filename
    filename = os.path.basename(excel_file).lower()
    if 'production' in filename:
        config_type = 'production'
        config = load_config('configs/production_config.json')
    elif 'warehouse' in filename:
        config_type = 'warehouse'
        config = load_config('configs/warehouse_config.json')
    elif 'idf' in filename:
        config_type = 'idf'
        config = load_config('configs/idf_config.json')
    else:
        raise ValueError(f"Unable to determine config type from filename: {filename}")
    
    # Filter data for the target month/year and config type
    df['Year/Month'] = df['Datetime'].dt.strftime('%Y-%m')
    target_year_month = target_date.strftime('%Y-%m')
    df_filtered = df[
        (df['Year/Month'] == target_year_month) & 
        (df['Task Type'].str.lower() == config_type)
    ]
    
    if len(df_filtered) == 0:
        print(f"No {config_type} tasks found for {target_date.strftime('%B %Y')} in the completion data")
        return
    
    # Load the existing Excel file
    try:
        wb = load_workbook(excel_file)
    except Exception as e:
        raise ValueError(f"Error loading Excel template: {os.path.basename(excel_file)}")
    
    # Create a case-insensitive mapping of sheet names
    sheet_name_mapping = {name.lower(): name for name in wb.sheetnames}
    
    # Track statistics for reporting
    total_records = len(df_filtered)
    matched_records = 0
    unmatched_records = []
    
    # Create column mappings for each sheet/frequency
    sheet_mappings = {}
    for section_name, frequencies in config.items():
        for freq_name, freq_config in frequencies.items():
            # Create mapping for this frequency
            column_mapping = {}
            for i, room in enumerate(freq_config['columns'], start=2):  # Start from column B
                column_mapping[normalize_string(room)] = i
            normalized_freq = normalize_string(freq_name)
            sheet_mappings[normalized_freq] = column_mapping
    
    # Process each completion record
    for _, row in df_filtered.iterrows():
        # Format the date string to match Excel format
        completion_date = row['Datetime'].strftime('%a, %b %-d %Y')
        
        # Get task frequency and room number (normalized)
        raw_freq = row['Frequency']
        task_freq = normalize_string(raw_freq)
        room_number = normalize_string(str(row['Room Number']))
        
        # Get the corresponding sheet name (case-insensitive lookup)
        actual_sheet_name = sheet_name_mapping.get(task_freq)
        if actual_sheet_name is None:
            unmatched_records.append(f"No sheet found for frequency '{raw_freq}'")
            continue
        
        ws = wb[actual_sheet_name]
        column_mapping = sheet_mappings.get(task_freq)
        
        if not column_mapping:
            unmatched_records.append(f"No column mapping found for frequency '{raw_freq}'")
            continue
        
        # Find the column for this room
        column = column_mapping.get(room_number)
        if column is None:
            unmatched_records.append(f"No matching column for room '{room_number}' in {raw_freq} sheet")
            continue
        
        # Find the row for this date
        date_cell = None
        for row_idx in range(3, ws.max_row + 1):  # Start from row 3 (after headers)
            if ws.cell(row=row_idx, column=1).value == completion_date:
                date_cell = row_idx
                break
        
        if date_cell is None:
            unmatched_records.append(f"Date not found: {completion_date}")
            continue
        
        # Add the completion data (full name)
        cell = ws.cell(row=date_cell, column=column)
        cell.value = row['Name']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Adjust column width if needed
        current_width = ws.column_dimensions[get_column_letter(column)].width
        name_width = len(row['Name']) + 2  # Add some padding
        if name_width > current_width:
            ws.column_dimensions[get_column_letter(column)].width = min(name_width, 20)  # Cap at 20 characters
        
        matched_records += 1
    
    # Save the workbook back to the same file
    try:
        wb.save(excel_file)
    except Exception as e:
        raise ValueError(f"Error saving workbook: {str(e)}")
    
    # Print summary
    print(f"\nPopulation Summary for {config_type.capitalize()} tasks - {target_date.strftime('%B %Y')}:")
    print(f"Total records processed: {total_records}")
    print(f"Successfully matched: {matched_records}")
    print(f"Unmatched records: {total_records - matched_records}")
    
    if unmatched_records:
        print("\nUnmatched Record Details:")
        for error in unmatched_records[:10]:  # Show first 10 errors
            print(f"- {error}")
        if len(unmatched_records) > 10:
            print(f"... and {len(unmatched_records) - 10} more errors")
        
        # Write unmatched tasks to review file
        write_unmatched_tasks(unmatched_records, excel_file, month, year)
    
    print(f"\nUpdated report saved: {os.path.basename(excel_file)}")

def find_latest_task_data(month: int, year: int) -> str:
    """
    Find the most recent task data file for the given month and year.
    
    Args:
        month (int): Month number (1-12)
        year (int): Year (e.g., 2025)
    
    Returns:
        str: Path to the most recent task data file
    """
    # Construct the pattern for the month/year
    pattern = f"task_data_{year}_{month:02d}_*.csv"
    processed_dir = os.path.join(SCRIPT_DIR, 'data', 'processed')
    
    # Get all matching files
    matching_files = []
    for file in os.listdir(processed_dir):
        if file.startswith(f"task_data_{year}_{month:02d}_") and file.endswith(".csv"):
            matching_files.append(os.path.join(processed_dir, file))
    
    if not matching_files:
        raise FileNotFoundError(f"No task data files found for {year}-{month:02d}")
    
    # Return the most recent file (based on filename timestamp)
    return max(matching_files)

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Generate task reports')
    parser.add_argument('--month', type=int, help='Month (1-12)')
    parser.add_argument('--year', type=int, help='Year (e.g., 2025)')
    parser.add_argument('--config', type=str, choices=['production', 'warehouse', 'idf', 'all'],
                       default='all', help='Which config to use')
    
    args = parser.parse_args()
    
    try:
        # Get month and year from command line or user input
        if args.month and args.year:
            if not (1 <= args.month <= 12):
                raise ValueError(f"Invalid month: {args.month}. Month must be between 1 and 12.")
            month = args.month
            year = args.year
        else:
            month, year = get_month_year_input()
        
        # Format month name for filename
        month_name = datetime.strptime(f"{month}/1/2000", "%m/%d/%Y").strftime("%b")
        
        # Create output directory for sign-off sheets
        output_dir = os.path.join(SCRIPT_DIR, 'data', 'signoff_sheets')
        os.makedirs(output_dir, exist_ok=True)
        
        # Find the latest task data file for the given month/year
        try:
            completion_data = find_latest_task_data(month, year)
            print(f"\nFound task data file: {os.path.basename(completion_data)}")
        except FileNotFoundError as e:
            print(f"\nWarning: {str(e)}")
            print("Will generate empty sign-off sheets.")
            completion_data = None
        
        # Generate reports based on config selection
        configs_to_run = []
        if args.config == 'all':
            configs_to_run = ['production', 'warehouse', 'idf']
        else:
            configs_to_run = [args.config]
        
        for config_name in configs_to_run:
            try:
                config = load_config(f'configs/{config_name}_config.json')
                output_filename = os.path.join(output_dir, f'{config_name.capitalize()}_Tasks_{month_name}{year}.xlsx')
                
                print(f"\nGenerating {output_filename}...")
                create_task_report(
                    month=month,
                    year=year,
                    output_filename=output_filename,
                    config=config
                )
                print(f"Created {output_filename}")
                
                # If completion data was found, populate it
                if completion_data:
                    print(f"\nPopulating {output_filename} with completion data...")
                    populate_report_data(output_filename, completion_data, month, year)
            
            except Exception as e:
                print(f"\nError processing {config_name} report: {str(e)}")
                print("Continuing with next report...")
                continue
    
    except Exception as e:
        print(f"\nError: {str(e)}")
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())